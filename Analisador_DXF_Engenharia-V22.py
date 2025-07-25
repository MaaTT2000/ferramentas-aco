import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import ezdxf
from ezdxf.math import Vec3
from ezdxf.addons.drawing import RenderContext, Frontend
from ezdxf.addons.drawing.matplotlib import MatplotlibBackend
from ezdxf.addons.drawing.properties import Properties

from collections import Counter
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import subprocess

# Importa a biblioteca de imagem Pillow (PIL)
from PIL import Image, ImageTk

# Imports para a visualização com Matplotlib
import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

# --- Funções de Análise ---

TOLERANCIA_GEOMETRIA = 1e-4

def get_flattend_entities(msp):
    """'Explode' todos os blocos para obter uma lista plana de entidades geométricas."""
    all_entities = []
    if msp is None:
        return all_entities
    for entity in msp:
        if entity.dxftype() == 'INSERT':
            try:
                # Usa virtual_entities para uma forma segura de explodir blocos
                all_entities.extend(entity.virtual_entities())
            except Exception as e:
                print(f"Não foi possível explodir o bloco: {e}")
                all_entities.append(entity)
        else:
            all_entities.append(entity)
    return all_entities

def analyze_layers(doc, msp):
    """Verifica se existem camadas vazias no desenho."""
    try:
        all_layers = doc.layers
        empty_layers = [layer.dxf.name for layer in all_layers if layer.dxf.name.lower() not in ['0', 'defpoints'] and not msp.query(f'*[layer=="{layer.dxf.name}"]')]
        if empty_layers: return {'status': 'ALERTA', 'details': f"Encontradas {len(empty_layers)} camadas vazias.", 'data': empty_layers}
        return {'status': 'OK', 'details': f'Nenhuma camada vazia encontrada ({len(all_layers)} total).', 'data': None}
    except Exception as e: return {'status': 'ERRO', 'details': f'Falha: {e}', 'data': None}

def summarize_entities(msp):
    """Conta e resume os tipos de entidades no desenho, incluindo as dentro de blocos."""
    try:
        all_entities = get_flattend_entities(msp)
        entity_counts = Counter(entity.dxftype() for entity in all_entities)
        if not entity_counts: return {'status': 'ALERTA', 'details': 'O desenho não possui elementos.', 'data': None}
        total_entities = sum(entity_counts.values())
        return {'status': 'OK', 'details': f"A peça é composta por {total_entities} elementos.", 'data': entity_counts}
    except Exception as e: return {'status': 'ERRO', 'details': f'Falha: {e}', 'data': None}

def find_duplicate_lines(msp, precision=4):
    """Encontra linhas duplicadas, incluindo as dentro de blocos."""
    try:
        all_entities = get_flattend_entities(msp)
        lines = [e for e in all_entities if e.dxftype() == 'LINE']
        duplicate_entities = set()
        for i in range(len(lines)):
            for j in range(i + 1, len(lines)):
                line1 = lines[i]
                line2 = lines[j]
                p1_start = tuple(round(c, precision) for c in line1.dxf.start)
                p1_end = tuple(round(c, precision) for c in line1.dxf.end)
                signature1 = tuple(sorted((p1_start, p1_end)))
                p2_start = tuple(round(c, precision) for c in line2.dxf.start)
                p2_end = tuple(round(c, precision) for c in line2.dxf.end)
                signature2 = tuple(sorted((p2_start, p2_end)))
                if signature1 == signature2:
                    duplicate_entities.add(line1)
                    duplicate_entities.add(line2)
        
        if duplicate_entities:
            return {'status': 'ERRO', 'details': f'Encontrada(s) {len(duplicate_entities)} linha(s) envolvida(s) em duplicação.', 'data': list(duplicate_entities)}
        
        return {'status': 'OK', 'details': 'Nenhuma linha sobreposta encontrada.', 'data': None}
    except Exception as e: return {'status': 'ERRO', 'details': f'Falha na análise de duplicatas: {e}', 'data': None}


def check_closed_geometry(msp, tolerance):
    """Verifica se o contorno da peça (linhas) está perfeitamente fechado."""
    try:
        all_entities = get_flattend_entities(msp)
        entities_to_check = [e for e in all_entities if e.dxftype() in ('LINE', 'LWPOLYLINE')]
        
        if not entities_to_check: return {'status': 'ALERTA', 'details': 'Nenhum contorno para analisar.', 'data': None}
        
        endpoints, open_points_coords = [], []
        def normalize_point(p): return tuple(round(coord / tolerance) for coord in p[:2])
        
        all_points = {}
        for entity in entities_to_check:
            if entity.dxftype() == 'LINE':
                start, end = entity.dxf.start, entity.dxf.end
                norm_start, norm_end = normalize_point(start), normalize_point(end)
                endpoints.extend([norm_start, norm_end])
                all_points[norm_start] = start
                all_points[norm_end] = end
        
        endpoint_counts = Counter(endpoints)
        open_normalized_points = {p for p, count in endpoint_counts.items() if count % 2 != 0}
        
        if open_normalized_points:
            open_points_coords = [all_points[p] for p in open_normalized_points]
            return {'status': 'ERRO', 'details': f'Contorno aberto! {len(open_points_coords)} pontos soltos.', 'data': open_points_coords}
        
        return {'status': 'OK', 'details': 'O contorno da peça está fechado.', 'data': None}
    except Exception as e: return {'status': 'ERRO', 'details': f'Falha: {e}', 'data': None}

def check_hole_symmetry(msp, tolerance=1e-2):
    """Verifica se o padrão de furos (círculos) é simétrico em relação ao centro da peça."""
    all_entities = get_flattend_entities(msp)
    holes = [e for e in all_entities if e.dxftype() == 'CIRCLE']
    if len(holes) < 2:
        return {'status': 'OK', 'details': 'Não há furos suficientes para análise de simetria.', 'data': None}

    bbox = DXFAnalyzerApp._calculate_manual_bbox(msp)
    if not bbox:
        return {'status': 'ALERTA', 'details': 'Não foi possível determinar o centro da peça.', 'data': None}
    
    min_x, max_x, min_y, max_y = bbox
    center_x = (min_x + max_x) / 2
    center_y = (min_y + max_y) / 2

    relative_positions = []
    for hole in holes:
        rel_x = hole.dxf.center.x - center_x
        rel_y = hole.dxf.center.y - center_y
        relative_positions.append({'pos': (rel_x, rel_y), 'entity': hole, 'matched': False})

    asymmetric_holes = []
    for i, p1 in enumerate(relative_positions):
        if p1['matched']:
            continue
        
        found_match = False
        for j, p2 in enumerate(relative_positions):
            if i == j or p2['matched']:
                continue
            
            is_symmetric_pair = abs(p1['pos'][0] + p2['pos'][0]) < tolerance and abs(p1['pos'][1] + p2['pos'][1]) < tolerance
            
            if is_symmetric_pair:
                p1['matched'] = True
                p2['matched'] = True
                found_match = True
                break
        
        is_at_center = abs(p1['pos'][0]) < tolerance and abs(p1['pos'][1]) < tolerance
        if not found_match and not is_at_center:
            asymmetric_holes.append(p1['entity'])

    if not asymmetric_holes:
        return {'status': 'OK', 'details': 'A furação da peça é simétrica.', 'data': None}
    else:
        return {'status': 'FURO SEM SIMETRIA', 'details': f'{len(asymmetric_holes)} furo(s) quebra(m) a simetria.', 'data': asymmetric_holes}

def check_drawing(filepath: Path):
    """Orquestra todas as verificações para um único arquivo DXF."""
    results = {}
    try:
        doc = ezdxf.readfile(filepath)
        msp = doc.modelspace()
    except Exception as e:
        results['load_error'] = f'Falha Crítica ao carregar: {e}'
        return results

    results['Composição do Desenho'] = summarize_entities(msp)
    results['Análise de Camadas'] = analyze_layers(doc, msp)
    results['Linhas Sobrepostas'] = find_duplicate_lines(msp)
    results['Verificação de Contorno'] = check_closed_geometry(msp, TOLERANCIA_GEOMETRIA)
    results['Verificação de Furos'] = check_hole_symmetry(msp)
    return results

class DXFAnalyzerApp:
    def _configure_styles(self):
        """Configura todos os estilos da aplicação em um só lugar."""
        self.style = ttk.Style(self.root)
        self.style.theme_use("clam")

        self.style.configure('.', background=self.BG_COLOR, foreground=self.FG_COLOR, fieldbackground=self.WIDGET_BG, borderwidth=1)
        self.style.configure('TFrame', background=self.BG_COLOR)
        self.style.configure('TLabel', background=self.BG_COLOR, foreground=self.FG_COLOR)
        
        self.style.configure('TButton', background=self.BUTTON_COLOR, foreground=self.FG_COLOR, borderwidth=0, focusthickness=0, padding=5)
        self.style.map('TButton', background=[('active', self.BUTTON_HOVER_COLOR)])
        
        self.style.configure('Accent.TButton', foreground=self.FG_COLOR, background=self.BUTTON_COLOR, padding=5)
        self.style.map('Accent.TButton', background=[('active', self.BUTTON_HOVER_COLOR)])

        self.style.configure("Treeview", background=self.WIDGET_BG, foreground=self.FG_COLOR, fieldbackground=self.WIDGET_BG, rowheight=28, font=('Segoe UI', 10))
        self.style.configure("Treeview.Heading", background=self.BG_COLOR, foreground=self.FG_COLOR, borderwidth=0, font=('Segoe UI', 10, 'bold'))
        self.style.map("Treeview", background=[('selected', self.SELECT_BG)], foreground=[('selected', self.FG_COLOR)])
        self.style.map("Treeview.Heading", background=[('active', self.BG_COLOR)])
        
        self.style.configure("Summary.TLabel", padding=5, relief="sunken", anchor="center", background=self.WIDGET_BG, font=('Segoe UI', 10))
        
        bold_font = ('Segoe UI', 9, 'bold')
        self.style.configure("Summary.TButton", padding=5, relief="flat", anchor="center", font=bold_font)
        
        self.style.configure("OK.TButton", background=self.OK_COLOR)
        self.style.map("OK.TButton", foreground=[('disabled', self.TEXT_ON_COLOR_BG)])
        
        self.style.configure("Error.TButton", background=self.ERROR_COLOR)
        self.style.map("Error.TButton", foreground=[('!disabled', self.TEXT_ON_COLOR_BG), ('disabled', self.TEXT_ON_COLOR_BG)])

        self.style.configure("Alert.TButton", background=self.ALERT_COLOR)
        self.style.map("Alert.TButton", foreground=[('!disabled', self.TEXT_ON_COLOR_BG), ('disabled', self.TEXT_ON_COLOR_BG)])

        self.style.configure("Symmetry.TButton", background=self.SYMMETRY_COLOR)
        self.style.map("Symmetry.TButton", foreground=[('!disabled', self.TEXT_ON_COLOR_BG), ('disabled', self.TEXT_ON_COLOR_BG)])


    def __init__(self, root):
        self.root = root
        self.root.title("Analisador de DXF v5.0 - INOVA PROCESS")
        self.root.geometry("1200x800")
        
        self.BG_COLOR = '#2E2E2E'
        self.FG_COLOR = '#EAEAEA'
        self.WIDGET_BG = '#3C3C3C'
        self.SELECT_BG = '#555555'
        self.BUTTON_COLOR = '#5A5A5A'
        self.BUTTON_HOVER_COLOR = '#6A6A6A'
        self.TEXT_ON_COLOR_BG = '#000000'

        self.OK_COLOR = '#4CAF50'
        self.ALERT_COLOR = '#FFC107'
        self.ERROR_COLOR = '#F44336'
        self.SYMMETRY_COLOR = '#2196F3'
        
        self.root.configure(bg=self.BG_COLOR)
        self._configure_styles()
        
        self.analysis_data = {}
        self.tree_item_to_path = {}
        
        top_frame = ttk.Frame(root, padding="10")
        top_frame.pack(side="top", fill="x")

        top_frame.columnconfigure(1, weight=1) 

        select_folder_button = ttk.Button(top_frame, text="Selecionar Pasta...", command=self.select_folder)
        select_folder_button.grid(row=0, column=0, sticky='w', padx=(0,10))
        
        self.folder_path = tk.StringVar()
        path_label = ttk.Label(top_frame, textvariable=self.folder_path, relief="sunken", padding=5, background=self.WIDGET_BG)
        path_label.grid(row=0, column=1, sticky='ew')

        right_actions_frame = ttk.Frame(top_frame)
        right_actions_frame.grid(row=0, column=2, sticky='e', padx=(10,0))
        
        try:
            img = Image.open("logo.png")
            img_resized = img.resize((40, 40), Image.Resampling.LANCZOS)
            self.logo_image = ImageTk.PhotoImage(img_resized)
            logo_label = ttk.Label(right_actions_frame, image=self.logo_image, background=self.BG_COLOR)
            logo_label.pack(side="right", padx=(10,0))
        except Exception as e:
            print(f"AVISO: Não foi possível carregar o logo. {e}")
        
        self.start_button = ttk.Button(right_actions_frame, text="Iniciar Análise", command=self.start_analysis_thread, style="Accent.TButton")
        self.start_button.pack(side="left")
        
        self.export_button = ttk.Button(right_actions_frame, text="Exportar Excel", command=self.export_to_excel, state="disabled")
        self.export_button.pack(side="left", padx=5)
        
        self.convert_button = ttk.Button(right_actions_frame, text="Converter DWG > DXF", command=self.convert_dwg_to_dxf)
        self.convert_button.pack(side="left")
        
        self.scroll_button = ttk.Button(right_actions_frame, text="↓ Rolar", command=self.scroll_tree_down)
        self.scroll_button.pack(side="left", padx=5)
        
        self.fix_contour_button = ttk.Button(right_actions_frame, text="Corrigir Contorno", command=self.fix_open_contour, state="disabled")
        self.fix_contour_button.pack(side="left", padx=5)

        self.oda_converter_path = None
        
        main_paned_window = ttk.PanedWindow(root, orient="horizontal")
        main_paned_window.pack(side="top", fill="both", expand=True, padx=10, pady=(5, 5))
        
        report_frame = ttk.Frame(main_paned_window)
        cols = ('status', 'details', 'mod_date')
        self.tree = ttk.Treeview(report_frame, columns=cols, show='tree headings')
        
        self.tree.heading('#0', text='Arquivo / Verificação')
        self.tree.column('#0', width=350, minwidth=250, anchor='w')
        self.tree.heading('status', text='Status')
        self.tree.column('status', width=150, anchor='w')
        self.tree.heading('details', text='Detalhes da Verificação')
        self.tree.column('details', width=350, anchor='w')
        self.tree.heading('mod_date', text='Data de Modificação')
        self.tree.column('mod_date', width=150, anchor='e')
        
        self.tree.pack(side='left', fill='both', expand=True)
        self.tree.bind('<<TreeviewSelect>>', self.on_tree_select)
        main_paned_window.add(report_frame, weight=2)
        
        self.tree.tag_configure('OK', foreground=self.FG_COLOR)
        self.tree.tag_configure('ALERTA', foreground=self.ALERT_COLOR)
        self.tree.tag_configure('ERRO', foreground=self.ERROR_COLOR)
        self.tree.tag_configure('FURO SEM SIMETRIA', foreground=self.SYMMETRY_COLOR)
        
        self.vis_frame = ttk.Frame(main_paned_window)
        main_paned_window.add(self.vis_frame, weight=3)
        
        self.canvas_widget = None
        self.toolbar = None
        
        ttk.Separator(root, orient='horizontal').pack(side="top", fill='x', padx=10)
        summary_frame = ttk.Frame(root, padding=(10, 5))
        summary_frame.pack(side="bottom", fill="x")

        summary_frame.columnconfigure(0, weight=2)
        summary_frame.columnconfigure(1, weight=1)
        summary_frame.columnconfigure(2, weight=1)
        summary_frame.columnconfigure(3, weight=1)
        summary_frame.columnconfigure(4, weight=1)

        self.total_var = tk.StringVar(value="Analisados: 0")
        self.last_error_iid, self.last_alert_iid, self.last_symmetry_iid = None, None, None

        ttk.Label(summary_frame, textvariable=self.total_var, style="Summary.TLabel").grid(row=0, column=0, sticky='ew', padx=2)
        
        self.ok_button = ttk.Button(summary_frame, text="✅ OK: 0", style="OK.TButton", state="disabled")
        self.ok_button.grid(row=0, column=1, sticky='ew', padx=2)
        
        self.error_button = ttk.Button(summary_frame, text="❌ Erros: 0", style="Error.TButton", command=lambda: self.find_next_by_status('ERRO', 'last_error_iid'), state="disabled")
        self.error_button.grid(row=0, column=2, sticky='ew', padx=2)
        
        self.alert_button = ttk.Button(summary_frame, text="⚠️ Atenção: 0", style="Alert.TButton", command=lambda: self.find_next_by_status('ALERTA', 'last_alert_iid'), state="disabled")
        self.alert_button.grid(row=0, column=3, sticky='ew', padx=2)
        
        self.symmetry_button = ttk.Button(summary_frame, text="🔵 Sem Simetria: 0", style="Symmetry.TButton", command=lambda: self.find_next_by_status('FURO SEM SIMETRIA', 'last_symmetry_iid'), state="disabled")
        self.symmetry_button.grid(row=0, column=4, sticky='ew', padx=2)
        
    def on_tree_select(self, event):
        selected_iid = self.tree.focus()
        if not selected_iid: return
        
        parent_iid = self.tree.parent(selected_iid) or selected_iid
        
        filepath = self.tree_item_to_path.get(parent_iid)
        if not filepath: return

        file_results = self.analysis_data.get(str(filepath))
        if not file_results: return
        
        has_contour_error = False
        for check, result in file_results.items():
            if check == 'Verificação de Contorno' and result['status'] == 'ERRO':
                has_contour_error = True
                break
        self.fix_contour_button.config(state="normal" if has_contour_error else "disabled")
        
        check_name = self.tree.item(selected_iid, "text").strip()
        if not self.tree.parent(selected_iid):
            check_name = ''
            highlight_data = None
        else:
            highlight_data = file_results.get(check_name, {}).get('data')
        
        self.visualize_dxf(filepath, highlight_data, check_name)

    @staticmethod
    def _calculate_manual_bbox(msp):
        all_entities = get_flattend_entities(msp)
        if not all_entities:
            return None
            
        min_x, max_x, min_y, max_y = float('inf'), float('-inf'), float('inf'), float('-inf')
        has_data = False
        for entity in all_entities:
            try:
                if entity.dxftype() in ('ARC', 'SPLINE', 'ELLIPSE'):
                    points = list(entity.flattening(distance=0.1))
                elif entity.dxftype() == 'LINE':
                    points = [entity.dxf.start, entity.dxf.end]
                elif entity.dxftype() in ('LWPOLYLINE', 'POLYLINE'):
                    points = list(entity.get_points('xy'))
                elif entity.dxftype() == 'CIRCLE':
                    center = entity.dxf.center
                    radius = entity.dxf.radius
                    points = [(center.x - radius, center.y - radius), (center.x + radius, center.y + radius)]
                else: continue
                has_data = True
                for p in points:
                    min_x, max_x, min_y, max_y = min(min_x, p[0]), max(max_x, p[0]), min(min_y, p[1]), max(max_y, p[1])
            except (AttributeError, TypeError, IndexError): continue
        return (min_x, max_x, min_y, max_y) if has_data else None

    def _get_geometric_properties(self, msp):
        props = {'width': 0, 'height': 0, 'extents': None}
        
        extents_data = self._calculate_manual_bbox(msp)
        if not extents_data: return props
            
        props['extents'] = extents_data
        min_x, max_x, min_y, max_y = extents_data
        props['width'] = max_x - min_x
        props['height'] = max_y - min_y
                
        return props

    def _get_hole_summary(self, msp, for_excel=False):
        try:
            all_entities = get_flattend_entities(msp)
            diameters = [round(entity.dxf.radius * 2, 2) for entity in all_entities if entity.dxftype() == 'CIRCLE']
            if not diameters:
                return "" if not for_excel else (0, "")
                
            diameter_counts = Counter(diameters)
            
            if for_excel:
                total_holes = sum(diameter_counts.values())
                diameter_str = ", ".join([f"{count}x{diameter}" for diameter, count in sorted(diameter_counts.items())])
                return total_holes, diameter_str

            summary_parts = [f"{count}x Ø{diameter:.2f}" for diameter, count in sorted(diameter_counts.items())]
            return "Furos: " + ", ".join(summary_parts)
        except:
            return "" if not for_excel else (0, "")

    def _adjust_plot_view(self, ax, geo_props, hole_summary_text, piece_name):
        title_lines = [f"Peça: {piece_name}"]
        
        title_lines.append(f"Largura: {geo_props.get('width', 0):.2f} | Altura: {geo_props.get('height', 0):.2f}")

        if hole_summary_text:
            title_lines.append(hole_summary_text)

        ax.set_title("\n".join(title_lines), fontsize=10, color=self.FG_COLOR, pad=20)
        
        extents_data = geo_props.get('extents')
        if extents_data:
            width = geo_props.get('width', 0)
            height = geo_props.get('height', 0)
            min_x, max_x, min_y, max_y = extents_data
            center_x, center_y = (min_x + max_x) / 2, (min_y + max_y) / 2
            view_size = max(width, height)
            margin = view_size * 0.1
            ax.set_xlim(center_x - (view_size / 2) - margin, center_x + (view_size / 2) + margin)
            ax.set_ylim(center_y - (view_size / 2) - margin, center_y + (view_size / 2) + margin)
        
        ax.set_aspect('equal', adjustable='box')
        ax.tick_params(colors=self.FG_COLOR)
        for spine in ax.spines.values():
            spine.set_edgecolor(self.FG_COLOR)

    def visualize_dxf(self, filepath, highlight_data, check_name):
        if self.canvas_widget: self.canvas_widget.destroy()
        if self.toolbar: self.toolbar.destroy()

        fig = Figure(figsize=(5, 4), dpi=100, facecolor=self.BG_COLOR)
        ax = fig.add_subplot(111, facecolor=self.WIDGET_BG)
        
        try:
            doc = ezdxf.readfile(filepath)
            msp = doc.modelspace()
            
            try: doc.layers.get('0').color = 7
            except: pass 
            
            geo_props = self._get_geometric_properties(msp)
            hole_summary = self._get_hole_summary(msp)
            
            backend = MatplotlibBackend(ax)
            Frontend(RenderContext(doc), backend).draw_layout(msp)
            
            all_entities = get_flattend_entities(msp)
            for entity in all_entities:
                if entity.dxftype() == 'CIRCLE':
                    center = entity.dxf.center
                    ax.text(center.x, center.y + entity.dxf.radius * 1.2, f"({center.x:.1f}, {center.y:.1f})",
                            color=self.FG_COLOR, fontsize=8, ha='center', va='bottom')

            if check_name == 'Linhas Sobrepostas' and highlight_data:
                error_props = Properties(color=self.ERROR_COLOR, lineweight=2.5)
                for line_entity in highlight_data:
                    try: backend.draw_line(line_entity.dxf.start, line_entity.dxf.end, properties=error_props)
                    except: continue

            if check_name == 'Verificação de Contorno' and highlight_data:
                x_coords, y_coords = [p[0] for p in highlight_data], [p[1] for p in highlight_data]
                ax.scatter(x_coords, y_coords, color=self.ERROR_COLOR, s=50, marker='o', facecolors='none', edgecolors=self.ERROR_COLOR, linewidths=1.5, zorder=5)
            
            if check_name == 'Verificação de Furos' and highlight_data:
                for hole_entity in highlight_data:
                    center, radius = hole_entity.dxf.center, hole_entity.dxf.radius
                    highlight_circle = plt.Circle((center.x, center.y), radius * 1.5, edgecolor=self.SYMMETRY_COLOR, facecolor='none', linewidth=1.5)
                    ax.add_artist(highlight_circle)

            self._adjust_plot_view(ax, geo_props, hole_summary, filepath.name)
            
            backend.finalize()

        except Exception as e:
            ax.text(0.5, 0.5, f"Erro ao visualizar o DXF:\n{e}", ha='center', va='center', wrap=True, color=self.FG_COLOR)
        
        canvas = FigureCanvasTkAgg(fig, master=self.vis_frame)
        self.canvas_widget = canvas.get_tk_widget()
        self.canvas_widget.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        self.toolbar = NavigationToolbar2Tk(canvas, self.vis_frame)
        self.toolbar.config(background=self.BG_COLOR)
        self.toolbar.update()
        
        canvas.draw()
            
    def start_analysis_thread(self):
        if not Path(self.folder_path.get()).is_dir():
            messagebox.showerror("Erro", "Por favor, selecione uma pasta válida.")
            return
        
        self.tree_item_to_path.clear()
        
        self.start_button.config(state="disabled")
        self.export_button.config(state="disabled")
        self.fix_contour_button.config(state="disabled")
        
        thread = threading.Thread(target=self.run_analysis_logic, daemon=True)
        thread.start()

    def run_analysis_logic(self):
        path = Path(self.folder_path.get())
        dxf_files = list(path.glob('[!~$]*.dxf'))

        if not dxf_files:
            self.root.after(0, lambda: messagebox.showinfo("Informação", "Nenhum arquivo .dxf encontrado na pasta selecionada."))
            self.root.after(0, lambda: self.start_button.config(state="normal"))
            return
            
        dxf_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
        
        all_results_for_ui = []
        for file_path in dxf_files:
            file_results = check_drawing(file_path)
            all_results_for_ui.append({'path': file_path, 'results': file_results})
        
        self.root.after(0, self.update_ui_with_results, all_results_for_ui)

    def update_ui_with_results(self, all_results):
        self.tree.delete(*self.tree.get_children())

        for data_packet in all_results:
            file_path = data_packet['path']
            file_results = data_packet['results']
            
            self.analysis_data[str(file_path)] = file_results
            
            mod_time = file_path.stat().st_mtime
            date_string = datetime.fromtimestamp(mod_time).strftime('%d/%m/%Y %H:%M')

            statuses = {res['status'] for res in file_results.values()}
            file_overall_status = 'OK'
            if 'ERRO' in statuses:
                file_overall_status = 'ERRO'
            elif 'FURO SEM SIMETRIA' in statuses:
                file_overall_status = 'FURO SEM SIMETRIA'
            elif 'ALERTA' in statuses:
                file_overall_status = 'ALERTA'

            should_be_open = (file_overall_status == 'ERRO')

            file_node = self.tree.insert('', 'end', text=f"📄 {file_path.name}", values=("", "", date_string), tags=(file_overall_status,), open=should_be_open)
            self.tree_item_to_path[file_node] = file_path
            
            if 'load_error' in file_results:
                self.tree.insert(file_node, 'end', text='  Erro de Carga', values=('❌ ERRO', file_results['load_error'], ""), tags=('ERRO',))
                continue

            for check_name, result in file_results.items():
                status = result['status']
                details = result['details']
                icon_map = {'OK': '✅', 'ALERTA': '⚠️', 'ERRO': '❌', 'FURO SEM SIMETRIA': '🔵'}
                icon = icon_map.get(status, '❓')
                self.tree.insert(file_node, 'end', text=f"  {check_name}", values=(f"{icon} {status}", details, ""), tags=(status,))
        
        self.update_summary_dashboard()
        self.start_button.config(state="normal")
        self.export_button.config(state="normal")

    def select_folder(self):
        path = filedialog.askdirectory(title="Selecione a pasta com os arquivos DXF")
        if path:
            self.folder_path.set(path)

    def export_to_excel(self):
        if not self.analysis_data:
            messagebox.showwarning("Atenção", "Nenhum dado de análise para exportar.")
            return
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Planilhas Excel", "*.xlsx"), ("Todos os Arquivos", "*.*")], title="Salvar Relatório")
        if not filepath: return
        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Relatório de Análise"
            headers = ["Nome da Peça", "Largura (mm)", "Altura (mm)", "Qtd. Furos", "Diâmetros", "Status Geral"]
            sheet.append(headers)
            header_font, center_align = Font(bold=True), Alignment(horizontal='center')
            for cell in sheet[1]: cell.font, cell.alignment = header_font, center_align
            for file_path_str, results in self.analysis_data.items():
                file_path, width, height, hole_qty, hole_dims_str, overall_status = Path(file_path_str), 0, 0, 0, "", "OK"
                
                status_flags = {res['status'] for res in results.values()}
                if 'ERRO' in status_flags: overall_status = 'ERRO'
                elif 'FURO SEM SIMETRIA' in status_flags: overall_status = 'FURO SEM SIMETRIA'
                elif 'ALERTA' in status_flags: overall_status = 'ALERTA'
                
                try:
                    doc = ezdxf.readfile(file_path); msp = doc.modelspace()
                    geo_props = self._get_geometric_properties(msp)
                    width, height = geo_props['width'], geo_props['height']
                    hole_qty, hole_dims_str = self._get_hole_summary(msp, for_excel=True)
                except Exception as e: print(f"Erro ao processar dados para Excel de {file_path.name}: {e}")

                sheet.append([file_path.name, f"{width:.2f}", f"{height:.2f}", hole_qty, hole_dims_str, overall_status])
            
            for col_idx, column_cells in enumerate(sheet.columns, 1):
                max_length = max(len(str(cell.value)) for cell in column_cells if cell.value) if any(c.value for c in column_cells) else 0
                sheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
            workbook.save(filepath)
            messagebox.showinfo("Sucesso", f"Relatório exportado para:\n{filepath}")
        except Exception as e: messagebox.showerror("Erro na Exportação", f"Ocorreu um erro:\n{e}")

    def _get_oda_converter_path(self):
        if self.oda_converter_path and Path(self.oda_converter_path).exists(): return True
        messagebox.showinfo("Configuração Necessária", "Aponte para o executável 'OdaFC.exe'.")
        path = filedialog.askopenfilename(title="Localize o OdaFC.exe", filetypes=[("ODA Converter", "OdaFC.exe"), ("Todos", "*.*")])
        if path: self.oda_converter_path = path; return True
        return False

    def convert_dwg_to_dxf(self):
        if not self._get_oda_converter_path(): return
        dwg_files = filedialog.askopenfilenames(title="Selecione arquivos DWG", filetypes=[("Arquivos DWG", "*.dwg"), ("Todos", "*.*")])
        if not dwg_files: return
        success_count, fail_count = 0, 0
        for file in dwg_files:
            file_path, output_folder = Path(file), Path(file).parent
            command = [self.oda_converter_path, str(file_path), str(output_folder), "ACAD2018", "DXF", "0", "1"]
            try:
                startupinfo = subprocess.STARTUPINFO(); startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
                result = subprocess.run(command, check=True, capture_output=True, text=True, startupinfo=startupinfo)
                if "conversion successful" in result.stdout.lower(): success_count += 1
                else: fail_count += 1; print(f"Falha em {file_path.name}:\n{result.stdout}\n{result.stderr}")
            except Exception as e: fail_count += 1; print(f"Erro ao executar conversor para {file_path.name}: {e}")
        if success_count > 0: messagebox.showinfo("Conversão Concluída", f"{success_count} arquivo(s) convertido(s). {fail_count} falha(s).")
        else: messagebox.showerror("Falha na Conversão", "Nenhum arquivo convertido. Verifique o console.")

    def update_summary_dashboard(self):
        total_files = len(self.analysis_data)
        ok_count, alert_count, error_count, symmetry_count = 0, 0, 0, 0
        for results in self.analysis_data.values():
            statuses = {res['status'] for res in results.values()}
            if 'ERRO' in statuses: error_count += 1
            elif 'FURO SEM SIMETRIA' in statuses: symmetry_count += 1
            elif 'ALERTA' in statuses: alert_count += 1
            else: ok_count += 1
        self.total_var.set(f"Analisados: {total_files}")
        self.ok_button.config(text=f"✅ OK: {ok_count}")
        self.error_button.config(text=f"❌ Erros: {error_count}", state="normal" if error_count > 0 else "disabled")
        self.alert_button.config(text=f"⚠️ Atenção: {alert_count}", state="normal" if alert_count > 0 else "disabled")
        self.symmetry_button.config(text=f"🔵 Sem Simetria: {symmetry_count}", state="normal" if symmetry_count > 0 else "disabled")
        self.last_error_iid, self.last_alert_iid, self.last_symmetry_iid = None, None, None

    def find_next_by_status(self, status_to_find, last_iid_attr):
        all_files = self.tree.get_children('')
        if not all_files: return
        last_iid = getattr(self, last_iid_attr, None)
        start_index = (all_files.index(last_iid) + 1) % len(all_files) if last_iid and last_iid in all_files else 0
        for i in range(len(all_files)):
            current_index = (start_index + i) % len(all_files)
            parent_iid = all_files[current_index]
            for child_iid in self.tree.get_children(parent_iid):
                if status_to_find in self.tree.item(child_iid, 'values')[0]:
                    self.tree.selection_set(parent_iid)
                    self.tree.focus(parent_iid)
                    self.tree.see(parent_iid)
                    setattr(self, last_iid_attr, parent_iid)
                    self.tree.event_generate("<<TreeviewSelect>>")
                    return

    def scroll_tree_down(self):
        if self.tree: self.tree.yview_scroll(5, "units")

    def fix_open_contour(self):
        """Tenta corrigir um contorno aberto unindo as linhas em uma LWPOLYLINE."""
        selected_iid = self.tree.focus()
        if not selected_iid:
            messagebox.showwarning("Ação Inválida", "Nenhum arquivo selecionado.")
            return

        parent_iid = self.tree.parent(selected_iid) or selected_iid
        filepath = self.tree_item_to_path.get(parent_iid)
        if not filepath:
            return

        try:
            doc = ezdxf.readfile(filepath)
            msp = doc.modelspace()

            all_entities = get_flattend_entities(msp)
            lines = [e for e in all_entities if e.dxftype() == 'LINE']
            
            if not lines:
                messagebox.showinfo("Informação", "Nenhuma linha encontrada para unir.")
                return

            # Algoritmo de reconstrução de caminho
            path_points = []
            remaining_lines = lines[:]
            current_line = remaining_lines.pop(0)
            path_points.extend([current_line.dxf.start, current_line.dxf.end])

            for _ in range(len(remaining_lines) * 2): # Limita as iterações para evitar loops infinitos
                last_point = Vec3(path_points[-1])
                found_next = False
                for i, next_line in enumerate(remaining_lines):
                    if last_point.isclose(next_line.dxf.start):
                        path_points.append(next_line.dxf.end)
                        remaining_lines.pop(i)
                        found_next = True
                        break
                    elif last_point.isclose(next_line.dxf.end):
                        path_points.append(next_line.dxf.start)
                        remaining_lines.pop(i)
                        found_next = True
                        break
                if not found_next:
                    break
            
            # Deleta as linhas antigas (apenas as do modelspace, não as virtuais)
            for line in msp.query('LINE'):
                msp.delete_entity(line)
            
            # Adiciona a nova polilinha fechada
            msp.add_lwpolyline(path_points, close=True)

            # Salva o arquivo com um novo nome
            new_filepath = filepath.with_name(f"{filepath.stem}_corrigido.dxf")
            doc.saveas(new_filepath)
            
            messagebox.showinfo("Sucesso", f"Contorno corrigido com sucesso!\nNovo arquivo salvo como:\n{new_filepath.name}")
            
            # Re-executa a análise para atualizar a lista
            self.start_analysis_thread()

        except Exception as e:
            messagebox.showerror("Erro na Correção", f"Não foi possível corrigir o contorno:\n{e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = DXFAnalyzerApp(root)
    root.mainloop()
